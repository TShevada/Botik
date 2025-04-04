import logging
import asyncio
import openpyxl
import os
import random
import string
from datetime import datetime
from aiogram import Bot, Dispatcher, types, F
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from collections import defaultdict
from aiohttp import web

# ===== CONFIGURATION =====
TOKEN = "7598421595:AAFIBwcEENiYq23qGLItJNGx6AHbAH7K17Y"
YOUR_TELEGRAM_ID = 1291104906
PAYMENT_CARD = "4169 7388 9268 3164"
# ========================

# Constants
PHOTOS_DIR = "payment_screenshots"
WELCOME_BANNER = "welcome_banner.jpg"
PORT = 10002

# Setup
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
os.makedirs(PHOTOS_DIR, exist_ok=True)

bot = Bot(token=TOKEN, parse_mode=ParseMode.MARKDOWN)
dp = Dispatcher()

# Storage (using dictionaries since we're not using FSM storage)
user_lang = {}
user_data = {}
save_counter = defaultdict(int)
admin_pending_actions = {}
pending_approvals = {}
ticket_codes = {}
# Ticket Prices - Updated structure
TICKET_TYPES = {
    "standard": {
        "ru": {
            "name": "Стандарт",
            "price": "20 AZN",
            "desc": "Приветственные коктейли, Fan Zone",
            "features": ["Приветственные коктейли", "Fan Zone"],
            "notice": "❗️Обратите внимание: после покупки билеты не подлежат возврату.❗️"
        },
        "az": {
            "name": "Standart",
            "price": "20 AZN",
            "desc": "Qarşılama kokteylləri, Fan Zone",
            "features": ["Qarşılama kokteylləri", "Fan Zone"],
            "notice": "❗️Diqqət: Biletləri aldıqdan sonra geri qaytarılmağa məcbur deyil.❗️"
        },
        "en": {
            "name": "Standard",
            "price": "20 AZN",
            "desc": "Welcome cocktails, Fan Zone",
            "features": ["Welcome cocktails", "Fan Zone"],
            "notice": "❗️Please note: tickets are non-refundable after purchase.❗️"
        }
    },
    "vip_single": {
        "ru": {
            "name": "VIP (Одиночный)",
            "price": "40 AZN",
            "desc": "Индивидуальное место за столиком, приветственный коктейль, количество мест ограничено",
            "features": ["Индивидуальное место за столиком", "Приветственный коктейль", "Количество мест ограничено"],
            "notice": "❗️Обратите внимание: после покупки билеты не подлежат возврату.❗️"
        },
        "az": {
            "name": "VIP (Tək)",
            "price": "40 AZN",
            "desc": "Fərdi oturacaq yeri, qarşılama kokteyli, məhdud sayda yer",
            "features": ["Fərdi oturacaq yeri", "Qarşılama kokteyli", "Məhdud sayda yer"],
            "notice": "❗️Diqqət: Biletləri aldıqdan sonra geri qaytarılmağa məcbur deyil.❗️"
        },
        "en": {
            "name": "VIP (Single)",
            "price": "40 AZN",
            "desc": "Individual seat at the table, welcome cocktail, limited space available",
            "features": ["Individual seat at the table", "Welcome cocktail", "Limited space available"],
            "notice": "❗️Please note: tickets are non-refundable after purchase.❗️"
        }
    },
    "vip_table": {
        "ru": {
            "name": "VIP (Столик)",
            "price": "160 AZN",
            "desc": "Отдельный столик на 4 человек, приветственные коктейли для всей компании, количество мест ограничено",
            "features": ["Отдельный столик на 4 человек", "Приветственные коктейли для всей компании", "Количество мест ограничено"],
            "notice": "❗️Обратите внимание: после покупки билеты не подлежат возврату.❗️"
        },
        "az": {
            "name": "VIP (Masalıq)",
            "price": "160 AZN",
            "desc": "4 nəfərlik ayrı masa, bütün şirkət üçün qarşılama kokteylləri, məhdud sayda yer",
            "features": ["4 nəfərlik ayrı masa", "Bütün şirkət üçün qarşılama kokteylləri", "Məhdud sayda yer"],
            "notice": "❗️Diqqət: Biletləri aldıqdan sonra geri qaytarılmağa məcbur deyil.❗️"
        },
        "en": {
            "name": "VIP (Table)",
            "price": "160 AZN",
            "desc": "Private table for 4 people, welcome cocktails for the whole group, limited space available",
            "features": ["Private table for 4 people", "Welcome cocktails for the whole group", "Limited space available"],
            "notice": "❗️Please note: tickets are non-refundable after purchase.❗️"
        }
    },
    "exclusive_table": {
        "ru": {
            "name": "Exclusive (Столик)",
            "price": "240 AZN",
            "desc": "Доступ прямо за DJ стол, отдельный столик на 4 человек, приветственные коктейли для всей компании, количество мест ограничено",
            "features": ["Доступ прямо за DJ стол", "Отдельный столик на 4 человек", "Приветственные коктейли для всей компании", "Количество мест ограничено"],
            "notice": "❗️Обратите внимание: после покупки билеты не подлежат возврату.❗️"
        },
        "az": {
            "name": "Exclusive (Masalıq)",
            "price": "240 AZN",
            "desc": "DJ masasının arxasına giriş, 4 nəfərlik ayrı masa, bütün şirkət üçün qarşılama kokteylləri, məhdud sayda yer",
            "features": ["DJ masasının arxasına giriş", "4 nəfərlik ayrı masa", "Bütün şirkət üçün qarşılama kokteylləri", "Məhdud sayda yer"],
            "notice": "❗️Diqqət: Biletləri aldıqdan sonra geri qaytarılmağa məcbur deyil.❗️"
        },
        "en": {
            "name": "Exclusive (Table)",
            "price": "240 AZN",
            "desc": "Access behind the DJ booth, private table for 4 people, welcome cocktails for the whole group, limited space available",
            "features": ["Access behind the DJ booth", "Private table for 4 people", "Welcome cocktails for the whole group", "Limited space available"],
            "notice": "❗️Please note: tickets are non-refundable after purchase.❗️"
        }
    }
}

# Helper Functions
def is_admin(user_id: int) -> bool:
    return user_id == YOUR_TELEGRAM_ID

def generate_ticket_code(user_id: int):
    """Generate a unique 8-digit alphanumeric ticket code"""
    code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    ticket_codes[user_id] = code
    return code

def get_lang_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🇷🇺 Русский"), KeyboardButton(text="🇦🇿 Azərbaycan")],
            [KeyboardButton(text="🇬🇧 English")]
        ],
        resize_keyboard=True
    )

def get_menu_keyboard(lang):
    if lang == "ru":
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="🎫 Билеты")],
                [KeyboardButton(text="📅 Ближайшие события")],
                [KeyboardButton(text="📞 Контакты")],
                [KeyboardButton(text="🌐 Сменить язык")]
            ],
            resize_keyboard=True
        )
    elif lang == "az":
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="🎫 Biletlər")],
                [KeyboardButton(text="📅 Yaxın tədbirlər")],
                [KeyboardButton(text="📞 Əlaqə")],
                [KeyboardButton(text="🌐 Dil dəyiş")]
            ],
            resize_keyboard=True
        )
    else:
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="🎫 Tickets")],
                [KeyboardButton(text="📅 Upcoming events")],
                [KeyboardButton(text="📞 Contacts")],
                [KeyboardButton(text="🌐 Change language")]
            ],
            resize_keyboard=True
        )

def get_ticket_type_keyboard(lang):
    if lang == "ru":
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Стандарт (20 AZN)")],
                [KeyboardButton(text="VIP Одиночный (40 AZN)")],
                [KeyboardButton(text="VIP Столик (160 AZN)")],
                [KeyboardButton(text="Exclusive Столик (240 AZN)")],
                [KeyboardButton(text="⬅️ Назад")]
            ],
            resize_keyboard=True
        )
    elif lang == "az":
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Standart (20 AZN)")],
                [KeyboardButton(text="VIP Tək (40 AZN)")],
                [KeyboardButton(text="VIP Masalıq (160 AZN)")],
                [KeyboardButton(text="Exclusive Masalıq (240 AZN)")],
                [KeyboardButton(text="⬅️ Geri")]
            ],
            resize_keyboard=True
        )
    else:
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Standard (20 AZN)")],
                [KeyboardButton(text="VIP Single (40 AZN)")],
                [KeyboardButton(text="VIP Table (160 AZN)")],
                [KeyboardButton(text="Exclusive Table (240 AZN)")],
                [KeyboardButton(text="⬅️ Back")]
            ],
            resize_keyboard=True
        )

def get_admin_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Статистика", callback_data="admin_stats"),
         InlineKeyboardButton(text="📝 Последние заявки", callback_data="admin_last_orders")],
        [InlineKeyboardButton(text="🔍 Поиск по ID", callback_data="admin_search"),
         InlineKeyboardButton(text="🔄 Обновить", callback_data="admin_refresh")],
        [InlineKeyboardButton(text="⏳ Ожидающие", callback_data="admin_pending")]
    ])

async def generate_stats_report():
    try:
        wb = openpyxl.load_workbook("tickets.xlsx")
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        
        total = len(rows) - 1
        if total <= 0:
            return "📭 Нет данных о заявках."
            
        types_count = defaultdict(int)
        for row in rows[1:]:
            types_count[row[3]] += 1
            
        report = (
            f"📈 *Статистика заявок*\n\n"
            f"• Всего заявок: {total}\n"
            f"• Стандарт: {types_count.get('standard', 0)}\n"
            f"• VIP Одиночный: {types_count.get('vip_single', 0)}\n"
            f"• VIP Столик: {types_count.get('vip_table', 0)}\n"
            f"• Эксклюзив: {types_count.get('exclusive_table', 0)}\n\n"
            f"Ожидают подтверждения: {len([x for x in pending_approvals.values() if x['approved'] is None])}\n"
            f"Последняя запись:\n"
            f"🕒 {rows[-1][6]}"
        )
        return report
    except Exception as e:
        logger.error(f"Stats error: {e}")
        return f"⚠️ Ошибка генерации отчёта: {e}"

async def get_last_orders(count=5):
    try:
        wb = openpyxl.load_workbook("tickets.xlsx")
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[-count:]
        
        if len(rows) == 0:
            return "📭 Нет заявок."
            
        report = "📋 *Последние заявки:*\n\n"
        for row in rows:
            ticket_type = row[3]
            if ticket_type == "vip_single":
                ticket_type = "VIP Одиночный"
            elif ticket_type == "vip_table":
                ticket_type = "VIP Столик"
            elif ticket_type == "standard":
                ticket_type = "Стандарт"
            elif ticket_type == "exclusive_table":
                ticket_type = "Эксклюзив"
                
            report += (
                f"🔹 *ID:* {row[0]}\n"
                f"👤 *{row[1]}*\n"
                f"📞 `{row[2]}`\n"
                f"🎟 {ticket_type} ({row[4]})\n"
                f"🕒 {row[6]}\n"
                "━━━━━━━━━━━━━━\n"
            )
        return report
    except Exception as e:
        logger.error(f"Orders error: {e}")
        return f"⚠️ Ошибка: {e}"

async def get_pending_orders():
    try:
        if not pending_approvals:
            return "⏳ Нет ожидающих заявок."
            
        report = "⏳ *Ожидающие заявки:*\n\n"
        for user_id, data in pending_approvals.items():
            if data['approved'] is None:
                ticket_type = data['data'][3]
                if ticket_type == "vip_single":
                    ticket_type = "VIP Одиночный"
                elif ticket_type == "vip_table":
                    ticket_type = "VIP Столик"
                elif ticket_type == "standard":
                    ticket_type = "Стандарт"
                elif ticket_type == "exclusive_table":
                    ticket_type = "Эксклюзив"
                    
                report += (
                    f"🔹 *ID:* {user_id}\n"
                    f"👤 *{data['data'][1]}*\n"
                    f"📞 `{data['data'][2]}`\n"
                    f"🎟 {ticket_type} ({data['data'][4]})\n"
                    f"🕒 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                    "━━━━━━━━━━━━━━\n"
                )
        return report
    except Exception as e:
        logger.error(f"Pending orders error: {e}")
        return f"⚠️ Ошибка: {e}"

def save_to_excel(user_id, name, phone, ticket_type, ticket_price, photo_path, ticket_code=None):
    try:
        file_path = "tickets.xlsx"
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["User ID", "Name", "Phone", "Ticket Type", "Ticket Price", "Photo Path", "Date", "Ticket Code"])
        
        ws.append([
            user_id,
            name,
            phone,
            ticket_type,
            ticket_price,
            photo_path,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ticket_code or "N/A"
        ])
        wb.save(file_path)
        
        save_counter['total'] += 1
        if save_counter['total'] % 10 == 0:
            backup_path = f"tickets_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(backup_path)
            logger.info(f"Created backup: {backup_path}")
        
        return True
    except Exception as e:
        logger.error(f"Database error: {e}")
        return False

async def notify_admin(user_id: int, name: str, phone: str, ticket_type: str, ticket_price: str):
    try:
        if not YOUR_TELEGRAM_ID:
            logger.error("Admin ID not set")
            return
            
        # Translate ticket type for admin notification
        display_type = ticket_type
        if ticket_type == "vip_single":
            display_type = "VIP Одиночный"
        elif ticket_type == "vip_table":
            display_type = "VIP Столик"
        elif ticket_type == "standard":
            display_type = "Стандарт"
        elif ticket_type == "exclusive_table":
            display_type = "Эксклюзив"
            
        msg = await bot.send_message(
            chat_id=YOUR_TELEGRAM_ID,
            text=(
                f"🆕 *Новая заявка на билет*\n\n"
                f"👤 ID: {user_id}\n"
                f"📛 Имя: {name}\n"
                f"📱 Телефон: `{phone}`\n"
                f"🎫 Тип: {display_type}\n"
                f"💵 Сумма: {ticket_price}\n"
                f"🕒 Время: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                f"Ответьте на это сообщение командой:\n"
                f"/accept - подтвердить\n"
                f"/reject [причина] - отклонить"
            ),
            parse_mode="Markdown"
        )
        
        pending_approvals[user_id] = {
            "message_id": msg.message_id,
            "data": (user_id, name, phone, ticket_type, ticket_price),
            "approved": None
        }
    except Exception as e:
        logger.error(f"Failed to notify admin: {e}")

# Handlers
@dp.message(Command("start"))
async def start_cmd(message: types.Message):
    try:
        if os.path.exists(WELCOME_BANNER):
            await message.answer_photo(types.InputFile(WELCOME_BANNER))
    except Exception as e:
        logger.error(f"Banner error: {e}")
    await message.answer("Выберите язык / Select language / Dil seçin:", reply_markup=get_lang_keyboard())

@dp.message(F.text.in_(["🇷🇺 Русский", "🇦🇿 Azərbaycan", "🇬🇧 English"]))
async def set_language(message: types.Message):
    lang_map = {
        "🇷🇺 Русский": "ru",
        "🇦🇿 Azərbaycan": "az",
        "🇬🇧 English": "en"
    }
    user_lang[message.from_user.id] = lang_map[message.text]
    
    confirmation = {
        "ru": "Язык установлен: Русский",
        "az": "Dil seçildi: Azərbaycan",
        "en": "Language set: English"
    }[lang_map[message.text]]
    
    await message.answer(confirmation, reply_markup=get_menu_keyboard(lang_map[message.text]))

@dp.message(F.text.in_(["📅 Ближайшие события", "📅 Yaxın tədbirlər", "📅 Upcoming events"]))
async def events_handler(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    events_info = {
        "ru": "Текущий ивент: Afro-Party в Voodoo!\n"
              "📅 Дата: 27 апреля 2025\n"
              "🕒 Время: 18:00 - 00:00\n"
              "📍 Место: Рестобар Voodoo, ТРЦ Наргиз Молл, 3 этаж",
        "az": "Cari tədbir: Afro-Party Voodoo-da!\n"
              "📅 Tarix: 27 Aprel 2025\n"
              "🕒 Vaxt: 18:00 - 00:00\n"
              "📍 Yer: Voodoo Restobar, Nargiz Mall, 3-cü mərtəbə",
        "en": "Current event: Afro-Party at Voodoo!\n"
              "📅 Date: April 27, 2025\n"
              "🕒 Time: 6:00 PM - 12:00 AM\n"
              "📍 Location: Voodoo Restobar, Nargiz Mall, 3rd floor"
    }[lang]
    await message.answer(events_info)

@dp.message(F.text.in_(["📞 Контакты", "📞 Əlaqə", "📞 Contacts"]))
async def contacts_handler(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    contact_info = {
        "ru": "📞 Контакты:\nТелефон: +994 10 531 24 06",
        "az": "📞 Əlaqə:\nTelefon: +994 10 531 24 06",
        "en": "📞 Contacts:\nPhone: +994 10 531 24 06"
    }[lang]
    await message.answer(contact_info)

@dp.message(F.text.in_(["🌐 Сменить язык", "🌐 Dil dəyiş", "🌐 Change language"]))
async def change_lang_handler(message: types.Message):
    await message.answer(
        "Выберите язык / Select language / Dil seçin:",
        reply_markup=get_lang_keyboard()
    )

@dp.message(F.text.in_(["🎫 Билеты", "🎫 Biletlər", "🎫 Tickets"]))
async def tickets_menu_handler(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    
    tickets_info = {
        "ru": "🎟 Доступные билеты:\n\n"
              f"1. {TICKET_TYPES['standard']['ru']['name']} - {TICKET_TYPES['standard']['ru']['price']}\n"
              f"   {TICKET_TYPES['standard']['ru']['desc']}\n\n"
              f"2. {TICKET_TYPES['vip_single']['ru']['name']} - {TICKET_TYPES['vip_single']['ru']['price']}\n"
              f"   {TICKET_TYPES['vip_single']['ru']['desc']}\n\n"
              f"3. {TICKET_TYPES['vip_table']['ru']['name']} - {TICKET_TYPES['vip_table']['ru']['price']}\n"
              f"   {TICKET_TYPES['vip_table']['ru']['desc']}\n\n"
              f"4. {TICKET_TYPES['exclusive_table']['ru']['name']} - {TICKET_TYPES['exclusive_table']['ru']['price']}\n"
              f"   {TICKET_TYPES['exclusive_table']['ru']['desc']}\n\n"
              f"{TICKET_TYPES['standard']['ru']['notice']}\n\n"
              "Выберите тип билета:",
        "az": "🎟 Mövcud biletlər:\n\n"
              f"1. {TICKET_TYPES['standard']['az']['name']} - {TICKET_TYPES['standard']['az']['price']}\n"
              f"   {TICKET_TYPES['standard']['az']['desc']}\n\n"
              f"2. {TICKET_TYPES['vip_single']['az']['name']} - {TICKET_TYPES['vip_single']['az']['price']}\n"
              f"   {TICKET_TYPES['vip_single']['az']['desc']}\n\n"
              f"3. {TICKET_TYPES['vip_table']['az']['name']} - {TICKET_TYPES['vip_table']['az']['price']}\n"
              f"   {TICKET_TYPES['vip_table']['az']['desc']}\n\n"
              f"4. {TICKET_TYPES['exclusive_table']['az']['name']} - {TICKET_TYPES['exclusive_table']['az']['price']}\n"
              f"   {TICKET_TYPES['exclusive_table']['az']['desc']}\n\n"
              f"{TICKET_TYPES['standard']['az']['notice']}\n\n"
              "Bilet növünü seçin:",
        "en": "🎟 Available tickets:\n\n"
              f"1. {TICKET_TYPES['standard']['en']['name']} - {TICKET_TYPES['standard']['en']['price']}\n"
              f"   {TICKET_TYPES['standard']['en']['desc']}\n\n"
              f"2. {TICKET_TYPES['vip_single']['en']['name']} - {TICKET_TYPES['vip_single']['en']['price']}\n"
              f"   {TICKET_TYPES['vip_single']['en']['desc']}\n\n"
              f"3. {TICKET_TYPES['vip_table']['en']['name']} - {TICKET_TYPES['vip_table']['en']['price']}\n"
              f"   {TICKET_TYPES['vip_table']['en']['desc']}\n\n"
              f"4. {TICKET_TYPES['exclusive_table']['en']['name']} - {TICKET_TYPES['exclusive_table']['en']['price']}\n"
              f"   {TICKET_TYPES['exclusive_table']['en']['desc']}\n\n"
              f"{TICKET_TYPES['standard']['en']['notice']}\n\n"
              "Select ticket type:"
    }[lang]
    
    await message.answer(tickets_info, reply_markup=get_ticket_type_keyboard(lang))

@dp.message(F.text.in_(["⬅️ Назад", "⬅️ Geri", "⬅️ Back"]))
async def back_handler(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    await message.answer("Главное меню" if lang == "ru" else "Ana menyu" if lang == "az" else "Main menu", 
                        reply_markup=get_menu_keyboard(lang))

@dp.message(F.text.regexp(r"(Стандарт|Standart|Standard|VIP.*|Exclusive.*|Эксклюзив|Eksklüziv).*"))
async def ticket_type_handler(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    
    ticket_type = None
    if "Стандарт" in message.text or "Standart" in message.text or "Standard" in message.text:
        ticket_type = "standard"
    elif "VIP Одиночный" in message.text or "VIP Tək" in message.text or "VIP Single" in message.text:
        ticket_type = "vip_single"
    elif "VIP Столик" in message.text or "VIP Masalıq" in message.text or "VIP Table" in message.text:
        ticket_type = "vip_table"
    elif "Exclusive" in message.text or "Эксклюзив" in message.text or "Eksklüziv" in message.text:
        ticket_type = "exclusive_table"
    
    if not ticket_type:
        await message.answer("Неверный тип билета" if lang == "ru" else "Yanlış bilet növü" if lang == "az" else "Invalid ticket type")
        return
    
    user_data[message.from_user.id] = {
        "step": "name",
        "lang": lang,
        "ticket_type": ticket_type,
        "ticket_price": TICKET_TYPES[ticket_type][lang]["price"],
        "name": None,
        "phone": None
    }
    
    prompt = {
        "ru": "Для покупки билетов введите ваше Имя и Фамилию:",
        "az": "Bilet almaq üçün ad və soyadınızı daxil edin:",
        "en": "To buy tickets, please enter your First and Last name:"
    }[lang]
    
    await message.answer(prompt, reply_markup=types.ReplyKeyboardRemove())

@dp.message(lambda m: user_data.get(m.from_user.id, {}).get("step") == "name")
async def get_name(message: types.Message):
    try:
        if message.from_user.id not in user_data:
            lang = user_lang.get(message.from_user.id, "en")
            await message.answer(
                "Пожалуйста, выберите тип билета сначала" if lang == "ru" else
                "Zəhmət olmasa, əvvəlcə bilet növünü seçin" if lang == "az" else
                "Please select ticket type first",
                reply_markup=get_menu_keyboard(lang)
            )
            return

        if len(message.text.split()) < 2:
            lang = user_data[message.from_user.id].get("lang", "en")
            error_msg = {
                "ru": "Пожалуйста, введите имя и фамилию",
                "az": "Zəhmət olmasa, ad və soyadınızı daxil edin",
                "en": "Please enter both first and last name"
            }[lang]
            await message.answer(error_msg)
            return

        user_data[message.from_user.id]["name"] = message.text
        user_data[message.from_user.id]["step"] = "phone"
        lang = user_data[message.from_user.id].get("lang", "en")
        
        prompt = {
            "ru": "Теперь введите ваш номер телефона:",
            "az": "İndi telefon nömrənizi daxil edin:",
            "en": "Now please enter your phone number:"
        }[lang]
        
        await message.answer(prompt)
        
    except Exception as e:
        logger.error(f"Error in get_name handler: {e}")
        lang = user_lang.get(message.from_user.id, "en")
        error_msg = {
            "ru": "Произошла ошибка, пожалуйста, попробуйте снова",
            "az": "Xəta baş verdi, zəhmət olmasa yenidən cəhd edin",
            "en": "An error occurred, please try again"
        }[lang]
        await message.answer(error_msg, reply_markup=get_menu_keyboard(lang))
        if message.from_user.id in user_data:
            del user_data[message.from_user.id]

@dp.message(lambda m: user_data.get(m.from_user.id, {}).get("step") == "phone")
async def get_phone(message: types.Message):
    try:
        if message.from_user.id not in user_data:
            lang = user_lang.get(message.from_user.id, "en")
            await message.answer("Пожалуйста, начните процесс заново" if lang == "ru" else 
                                "Zəhmət olmasa, prosesi yenidən başladın" if lang == "az" else 
                                "Please start the process again")
            return

        phone = message.text
        if not phone.replace('+', '').isdigit() or len(phone.replace('+', '')) < 9:
            lang = user_data[message.from_user.id].get("lang", "en")
            error_msg = {
                "ru": "Пожалуйста, введите корректный номер телефона",
                "az": "Zəhmət olmasa, düzgün telefon nömrəsi daxil edin",
                "en": "Please enter a valid phone number"
            }[lang]
            await message.answer(error_msg)
            return
        
        user_data[message.from_user.id]["phone"] = phone
        user_data[message.from_user.id]["step"] = "confirm"
        lang = user_data[message.from_user.id].get("lang", "en")
        
        ticket_type = user_data[message.from_user.id]["ticket_type"]
        ticket_info = TICKET_TYPES[ticket_type][lang]
        
        confirmation = {
            "ru": f"Проверьте ваши данные:\n\n"
                  f"🎟 Тип билета: {ticket_info['name']}\n"
                  f"💳 Сумма: {ticket_info['price']}\n"
                  f"👤 Имя: {user_data[message.from_user.id]['name']}\n"
                  f"📱 Телефон: {phone}\n\n"
                  f"Все верно?",
            "az": f"Məlumatlarınızı yoxlayın:\n\n"
                  f"🎟 Bilet növü: {ticket_info['name']}\n"
                  f"💳 Məbləğ: {ticket_info['price']}\n"
                  f"👤 Ad: {user_data[message.from_user.id]['name']}\n"
                  f"📱 Telefon: {phone}\n\n"
                  f"Hər şey düzgündür?",
            "en": f"Please confirm your details:\n\n"
                  f"🎟 Ticket type: {ticket_info['name']}\n"
                  f"💳 Amount: {ticket_info['price']}\n"
                  f"👤 Name: {user_data[message.from_user.id]['name']}\n"
                  f"📱 Phone: {phone}\n\n"
                  f"Is everything correct?"
        }[lang]
        
        keyboard = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="✅ Да" if lang == "ru" else "✅ Bəli" if lang == "az" else "✅ Yes")],
                [KeyboardButton(text="❌ Нет" if lang == "ru" else "❌ Xeyr" if lang == "az" else "❌ No")]
            ],
            resize_keyboard=True
        )
        
        await message.answer(confirmation, reply_markup=keyboard)
    except Exception as e:
        logger.error(f"Error in get_phone handler: {e}")
        lang = user_lang.get(message.from_user.id, "en")
        error_msg = {
            "ru": "Произошла ошибка, пожалуйста, начните заново",
            "az": "Xəta baş verdi, zəhmət olmasa yenidən başlayın",
            "en": "An error occurred, please start over"
        }[lang]
        await message.answer(error_msg, reply_markup=get_menu_keyboard(lang))
        if message.from_user.id in user_data:
            del user_data[message.from_user.id]

@dp.message(F.text.in_(["✅ Да", "✅ Bəli", "✅ Yes"]))
async def confirm_purchase(message: types.Message):
    if message.from_user.id not in user_data:
        lang = user_lang.get(message.from_user.id, "en")
        await message.answer("Пожалуйста, начните процесс заново" if lang == "ru" else 
                            "Zəhmət olmasa, prosesi yenidən başladın" if lang == "az" else 
                            "Please start the process again")
        return
    
    lang = user_data[message.from_user.id].get("lang", "en")
    user_data[message.from_user.id]["step"] = "payment"
    
    payment_info = {
        "ru": f"Оплатите {user_data[message.from_user.id]['ticket_price']} на карту: `{PAYMENT_CARD}`\n"
              "и отправьте скриншот оплаты.\n\n"
              f"{TICKET_TYPES[user_data[message.from_user.id]['ticket_type']]['ru']['notice']}",
        "az": f"{user_data[message.from_user.id]['ticket_price']} məbləğini kartla ödəyin: `{PAYMENT_CARD}`\n"
              "və ödəniş skrinşotu göndərin.\n\n"
              f"{TICKET_TYPES[user_data[message.from_user.id]['ticket_type']]['az']['notice']}",
        "en": f"Please pay {user_data[message.from_user.id]['ticket_price']} to card: `{PAYMENT_CARD}`\n"
              "and send payment screenshot.\n\n"
              f"{TICKET_TYPES[user_data[message.from_user.id]['ticket_type']]['en']['notice']}"
    }[lang]
    
    await message.answer(payment_info, reply_markup=get_menu_keyboard(lang))

@dp.message(F.text.in_(["❌ Нет", "❌ Xeyr", "❌ No"]))
async def cancel_purchase(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    if message.from_user.id in user_data:
        del user_data[message.from_user.id]
    
    msg = {
        "ru": "Заказ отменен. Можете начать заново.",
        "az": "Sifariş ləğv edildi. Yenidən başlaya bilərsiniz.",
        "en": "Order canceled. You can start again."
    }[lang]
    
    await message.answer(msg, reply_markup=get_menu_keyboard(lang))

@dp.message(lambda m: user_data.get(m.from_user.id, {}).get("step") == "payment")
async def handle_payment(message: types.Message):
    lang = user_data[message.from_user.id].get("lang", "en")
    
    if message.photo:
        try:
            photo = message.photo[-1]
            file = await bot.get_file(photo.file_id)
            path = f"{PHOTOS_DIR}/{message.from_user.id}_{photo.file_id}.jpg"
            await bot.download_file(file.file_path, path)
            
            # Generate unique ticket code
            ticket_code = generate_ticket_code(message.from_user.id)
            
            if save_to_excel(
                message.from_user.id,
                user_data[message.from_user.id]["name"],
                user_data[message.from_user.id]["phone"],
                user_data[message.from_user.id]["ticket_type"],
                user_data[message.from_user.id]["ticket_price"],
                path,
                ticket_code
            ):
                await notify_admin(
                    message.from_user.id,
                    user_data[message.from_user.id]["name"],
                    user_data[message.from_user.id]["phone"],
                    user_data[message.from_user.id]["ticket_type"],
                    user_data[message.from_user.id]["ticket_price"]
                )
                
                # Send ticket code to user
                ticket_type = user_data[message.from_user.id]["ticket_type"]
                ticket_info = TICKET_TYPES[ticket_type][lang]
                
                confirmation = {
                    "ru": f"Спасибо! Ваша заявка на рассмотрении.\n\n"
                          f"🔢 Ваш код билета: `{ticket_code}`\n"
                          f"🎫 Тип билета: {ticket_info['name']}\n"
                          f"👤 Имя: {user_data[message.from_user.id]['name']}\n\n"
                          f"Сохраните этот код для входа на мероприятие.",
                    "az": f"Təşəkkürlər! Müraciətiniz nəzərdən keçirilir.\n\n"
                          f"🔢 Bilet kodunuz: `{ticket_code}`\n"
                          f"🎫 Bilet növü: {ticket_info['name']}\n"
                          f"👤 Ad: {user_data[message.from_user.id]['name']}\n\n"
                          f"Tədbirə giriş üçün bu kodu saxlayın.",
                    "en": f"Thank you! Your application is under review.\n\n"
                          f"🔢 Your ticket code: `{ticket_code}`\n"
                          f"🎫 Ticket type: {ticket_info['name']}\n"
                          f"👤 Name: {user_data[message.from_user.id]['name']}\n\n"
                          f"Keep this code for event entry."
                }[lang]
                
                await message.answer(confirmation, reply_markup=get_menu_keyboard(lang))
                del user_data[message.from_user.id]
            
        except Exception as e:
            logger.error(f"Payment processing error: {e}")
            error_msg = {
                "ru": "Ошибка обработки платежа, попробуйте снова",
                "az": "Ödəniş emalı xətası, yenidən cəhd edin",
                "en": "Payment processing error, please try again"
            }[lang]
            await message.answer(error_msg)
    else:
        prompt = {
            "ru": "Пожалуйста, отправьте скриншот оплаты.",
            "az": "Zəhmət olmasa, ödəniş skrinşotu göndərin.",
            "en": "Please send the payment screenshot."
        }[lang]
        await message.answer(prompt)

@dp.message(Command("admin"))
async def admin_command(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Доступ запрещён!")
        return
        
    await message.answer(
        "🛠 *Панель администратора*",
        reply_markup=get_admin_keyboard(),
        parse_mode="Markdown"
    )

@dp.callback_query(F.data.startswith("admin_"))
async def handle_admin_callbacks(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Доступ запрещён!")
        return
    
    try:
        action = callback.data.split('_')[1]
        
        if action == "stats":
            report = await generate_stats_report()
            await callback.message.edit_text(report, reply_markup=get_admin_keyboard())
            
        elif action == "last_orders":
            orders = await get_last_orders()
            await callback.message.edit_text(orders, reply_markup=get_admin_keyboard())
            
        elif action == "search":
            await callback.message.answer("Введите ID пользователя:")
            admin_pending_actions[callback.from_user.id] = "waiting_for_id"
            
        elif action == "refresh":
            await callback.message.edit_text(
                "🛠 *Панель администратора*",
                reply_markup=get_admin_keyboard(),
                parse_mode="Markdown"
            )
            
        elif action == "pending":
            pending = await get_pending_orders()
            await callback.message.edit_text(pending, reply_markup=get_admin_keyboard())
            
        await callback.answer()
    except Exception as e:
        logger.error(f"Admin callback error: {e}")
        await callback.answer("⚠️ Произошла ошибка")
        
@dp.message(lambda m: admin_pending_actions.get(m.from_user.id) == "waiting_for_id")
async def handle_admin_search(message: types.Message):
    if not is_admin(message.from_user.id):
        return
        
    try:
        user_id = int(message.text)
        wb = openpyxl.load_workbook("tickets.xlsx")
        ws = wb.active
        
        found = None
        for row in ws.iter_rows(values_only=True):
            if row[0] == user_id:
                found = row
                break
                
        if not found:
            await message.answer("❌ Заявка не найдена")
        else:
            # Translate ticket type for display
            ticket_type = found[3]
            if ticket_type == "vip_single":
                ticket_type = "VIP Одиночный"
            elif ticket_type == "vip_table":
                ticket_type = "VIP Столик"
            elif ticket_type == "standard":
                ticket_type = "Стандарт"
            elif ticket_type == "exclusive_table":
                ticket_type = "Эксклюзив"
                
            report = (
                f"🔍 *Найдена заявка:*\n\n"
                f"👤 *{found[1]}*\n"
                f"📞 `{found[2]}`\n"
                f"🎟 {ticket_type} ({found[4]})\n"
                f"🔢 Код: `{found[7] if len(found) > 7 else 'N/A'}`\n"
                f"📸 [Фото]({found[5]})\n"
                f"🕒 {found[6]}"
            )
            await message.answer(report, parse_mode="Markdown")
            
    except ValueError:
        await message.answer("❌ Введите числовой ID")
    except Exception as e:
        logger.error(f"Search error: {e}")
        await message.answer("⚠️ Ошибка поиска")
    finally:
        admin_pending_actions.pop(message.from_user.id, None)

@dp.message(Command("accept"))
async def accept_request(message: types.Message):
    if not is_admin(message.from_user.id):
        return
        
    if not message.reply_to_message:
        await message.answer("ℹ️ Ответьте на сообщение с заявкой для подтверждения")
        return
        
    try:
        text = message.reply_to_message.text
        user_id = int(text.split("ID:")[1].split("\n")[0].strip())
        
        if user_id in pending_approvals:
            pending_approvals[user_id]["approved"] = True
            
            # Get ticket code
            ticket_code = ticket_codes.get(user_id, "N/A")
            
            # Send approval notification to user
            lang = user_lang.get(user_id, "en")
            approval_msg = {
                "ru": f"🎉 Ваша заявка подтверждена! Билет активен.\n\n"
                      f"🔢 Ваш код билета: `{ticket_code}`\n"
                      f"Сохраните этот код для входа на мероприятие.",
                "az": f"🎉 Müraciətiniz təsdiqləndi! Bilet aktivdir.\n\n"
                      f"🔢 Bilet kodunuz: `{ticket_code}`\n"
                      f"Tədbirə giriş üçün bu kodu saxlayın.",
                "en": f"🎉 Your application has been approved! Ticket is active.\n\n"
                      f"🔢 Your ticket code: `{ticket_code}`\n"
                      f"Keep this code for event entry."
            }[lang]
            
            await bot.send_message(user_id, approval_msg)
            await message.answer(f"✅ Заявка {user_id} подтверждена. Код билета: `{ticket_code}`")
            
            # Remove from pending
            del pending_approvals[user_id]
        else:
            await message.answer("⚠️ Заявка не найдена в ожидающих")
    except Exception as e:
        logger.error(f"Accept error: {e}")
        await message.answer("❌ Ошибка подтверждения")

@dp.message(Command("reject"))
async def reject_request(message: types.Message):
    if not is_admin(message.from_user.id):
        return
        
    if not message.reply_to_message:
        await message.answer("ℹ️ Ответьте на сообщение с заявкой для отклонения")
        return
        
    try:
        text = message.reply_to_message.text
        user_id = int(text.split("ID:")[1].split("\n")[0].strip())
        reason = message.text.split("/reject")[1].strip() or "не указана"
        
        if user_id in pending_approvals:
            pending_approvals[user_id]["approved"] = False
            
            # Send rejection notification to user
            lang = user_lang.get(user_id, "en")
            rejection_msg = {
                "ru": f"⚠️ Ваша заявка отклонена. Причина: {reason}",
                "az": f"⚠️ Müraciətiniz rədd edildi. Səbəb: {reason}",
                "en": f"⚠️ Your application was rejected. Reason: {reason}"
            }[lang]
            
            await bot.send_message(user_id, rejection_msg)
            await message.answer(f"❌ Заявка {user_id} отклонена. Причина: {reason}")
            
            # Remove from pending
            del pending_approvals[user_id]
        else:
            await message.answer("⚠️ Заявка не найдена в ожидающих")
    except Exception as e:
        logger.error(f"Reject error: {e}")
        await message.answer("❌ Ошибка отклонения")

@dp.message()
async def handle_unmatched_messages(message: types.Message):
    if message.from_user.id == YOUR_TELEGRAM_ID:
        await message.answer("ℹ️ Используйте /admin для управления ботом")
    else:
        # First check if user is in the middle of a process
        if message.from_user.id in user_data:
            current_step = user_data[message.from_user.id].get("step")
            if current_step == "name":
                await get_name(message)
                return
            elif current_step == "phone":
                await get_phone(message)
                return
            elif current_step == "payment":
                await handle_payment(message)
                return
        
        # Default response if not in any flow
        lang = user_lang.get(message.from_user.id, "en")
        response = {
            "ru": "Пожалуйста, используйте кнопки меню",
            "az": "Zəhmət olmasa menyu düymələrindən istifadə edin",
            "en": "Please use the menu buttons"
        }[lang]
        await message.answer(response, reply_markup=get_menu_keyboard(lang))
async def on_startup():
    await bot.send_message(YOUR_TELEGRAM_ID, "🤖 Bot started successfully!")

async def run_bot():
    await dp.start_polling(bot)

async def http_handler(request):
    return web.Response(text="🤖 Bot is running!")

async def main():
    # Start bot in background
    await on_startup()
    bot_task = asyncio.create_task(run_bot())

    # Configure HTTP server for Render
    app = web.Application()
    app.router.add_get("/", http_handler)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()

    logger.info(f"🚀 Bot running on port {PORT}")
    await asyncio.Event().wait()  # Run forever

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Bot stopped")
    except Exception as e:
        logger.critical(f"Fatal error: {e}")
